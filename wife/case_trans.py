import openpyxl
from openpyxl.utils import get_column_letter
import os
import re
import copy
from datetime import datetime
from openpyxl.styles import Color 

# 字段映射：源字段 -> 目标字段（用于定位单元格）
FIELD_MAP = {
    '报案号': '报案号',
    '被保人': '被保险人',
    '报案日期': '报案日期',
    '追偿首次发起日期': '追偿发起日期',
    '诉讼届满日期': '诉讼时效届满日',
    '追偿案件类型': '追偿类型',
    '被追偿人名称': '被追偿人',
    '计划追偿金额': '计划追偿金额',
    '案情简介': '追偿过程简介'
}

# 目标单元格坐标（基于模板的固定位置）
TARGET_CELLS = {
    '报案号': 'C4',
    '被保险人': 'C2',
    '报案日期': 'E4',
    '追偿发起日期': 'E8',
    '诉讼时效届满日': 'E9',
    '追偿类型': 'C8',
    '被追偿人': 'C9',
    '计划追偿金额': 'C10',
    '追偿过程简介': 'C13'
}


def sanitize_sheet_name(name):
    """清洗工作表名称：移除非法字符，截断长度"""
    if not name:
        name = "无名称"
    name = re.sub(r'[\\/*?:\[\]]', '_', name)
    return name[:31]


def read_source_data(source_path):
    """读取数据源文件，返回行字典列表"""
    wb = openpyxl.load_workbook(source_path, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    col_index = {header: idx + 1 for idx, header in enumerate(headers) if header}

    data_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for header, idx in col_index.items():
            row_dict[header] = row[idx - 1]
        data_rows.append(row_dict)
    return data_rows


def get_template_dimensions(template_ws):
    """从模板工作表获取所有自定义行高和列宽"""
    row_heights = {}
    for row_idx, row_dim in template_ws.row_dimensions.items():
        if row_dim.height is not None:
            row_heights[row_idx] = row_dim.height

    col_widths = {}
    for col_idx, col_dim in template_ws.column_dimensions.items():
        if col_dim.width is not None:
            # col_idx 可能是整数或字母，统一转为字母
            if isinstance(col_idx, int):
                col_letter = get_column_letter(col_idx)
            else:
                col_letter = col_idx
            col_widths[col_letter] = col_dim.width
    return row_heights, col_widths


def apply_template_customizations(sheet):
    """
    对工作表应用额外定制：
      - 合并指定区域
      - 设置放弃追偿金额公式
    """
    # 原有合并区域
    sheet.merge_cells('A1:E1')
    sheet.merge_cells('A2:A3')
    sheet.merge_cells('A4:A7')
    sheet.merge_cells('A8:A14')

    # 新增合并区域
    sheet.merge_cells('A15:B15')  # 合并A15和B15
    sheet.merge_cells('C6:E6')  # 合并C13、D13、E13（追偿过程简介区域）
    sheet.merge_cells('C7:E7')  # 合并C13、D13、E13（追偿过程简介区域）
    sheet.merge_cells('C12:E12')  # 合并C13、D13、E13（追偿过程简介区域）
    sheet.merge_cells('C13:E13')  # 合并C13、D13、E13（追偿过程简介区域）
    sheet.merge_cells('C14:E14')  # 合并C13、D13、E13（追偿过程简介区域）

    # 设置放弃追偿金额右侧单元格公式
    sheet['E11'] = '=C10-C11'
    sheet['C15'] = "舒媛"
    sheet['E15'] = datetime.now().strftime('%Y/%m/%d')

def fill_sheet(sheet, data_row):
    """根据数据行填充目标 sheet 的对应单元格"""
    for src_field, target_field in FIELD_MAP.items():
        if target_field in TARGET_CELLS:
            coord = TARGET_CELLS[target_field]
            value = data_row.get(src_field, '')
            if value is None:
                value = ''
            sheet[coord] = value


def create_output_file(template_path, data_group, group_index):
    """根据一组数据生成一个输出文件（最多20条）"""
    # 加载模板，获取模板工作表
    template_wb = openpyxl.load_workbook(template_path)
    template_ws = template_wb.active

    # 获取模板的行高和列宽
    row_heights, col_widths = get_template_dimensions(template_ws)

    # 创建输出工作簿
    output_wb = openpyxl.Workbook()
    output_wb.remove(output_wb.active)

    # 复制模板工作表作为空白模板（包含单元格值和样式）
    template_copy = output_wb.create_sheet("_template")
    for row in template_ws.iter_rows():
        for cell in row:
            new_cell = template_copy.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy.copy(cell.font)
                # 设置字体颜色为黑色
                from openpyxl.styles.colors import COLOR_INDEX
                new_font = copy.copy(new_cell.font)
                new_font.color = Color('FF000000')  # 黑色
                new_cell.font = new_font
                new_cell.border = copy.copy(cell.border)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy.copy(cell.protection)
                new_cell.alignment = copy.copy(cell.alignment)

    used_names = set()
    for data_row in data_group:
        raw_name = data_row.get('被追偿人名称', '')
        sheet_name = sanitize_sheet_name(raw_name)

        if sheet_name in used_names:
            suffix = 1
            while f"{sheet_name}_{suffix}" in used_names:
                suffix += 1
            sheet_name = f"{sheet_name}_{suffix}"
        used_names.add(sheet_name)

        # 复制模板并命名
        new_sheet = output_wb.copy_worksheet(output_wb["_template"])
        new_sheet.title = sheet_name

        # 应用模板的行高和列宽
        for row_idx, height in row_heights.items():
            new_sheet.row_dimensions[row_idx].height = height
        for col_letter, width in col_widths.items():
            new_sheet.column_dimensions[col_letter].width = width

        # 填充数据
        fill_sheet(new_sheet, data_row)
        # 应用额外定制（合并、公式）
        apply_template_customizations(new_sheet)

    # 删除模板工作表
    output_wb.remove(output_wb["_template"])

    output_filename = f"目标表{group_index}.xlsx"
    output_wb.save(output_filename)
    print(f"已生成: {output_filename}")


def main():
    source_file = "数据源.xlsx"
    template_file = "目标表.xlsx"

    if not os.path.exists(source_file):
        print(f"错误：找不到数据源文件 {source_file}")
        return
    if not os.path.exists(template_file):
        print(f"错误：找不到模板文件 {template_file}")
        return

    data_rows = read_source_data(source_file)
    print(f"共读取 {len(data_rows)} 条数据")

    group_size = 20
    groups = [data_rows[i:i + group_size] for i in range(0, len(data_rows), group_size)]

    for idx, group in enumerate(groups, start=1):
        create_output_file(template_file, group, idx)

    print("处理完成！")


if __name__ == "__main__":
    main()